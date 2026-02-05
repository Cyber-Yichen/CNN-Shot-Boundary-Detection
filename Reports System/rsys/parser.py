import os
from zipfile import ZipFile
from openpyxl import load_workbook
from .utils import safe_float, safe_int
from .config import SHEET_PREVIEW_LIMIT


def parse_kv_sheet(ws):
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        k = row[0]
        if k is None:
            continue
        k = str(k).strip()
        if not k:
            continue
        kl = k.lower()
        if kl in ("key", "metric", "name"):
            continue
        v = row[1] if len(row) > 1 else None
        out[kl] = v
    return out


def sheet_to_table(ws, limit_rows=SHEET_PREVIEW_LIMIT):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"headers": [], "rows": [], "total_rows": 0, "shown_rows": 0}

    headers = ["" if c is None else str(c) for c in rows[0]]
    body = []
    for r in rows[1 : 1 + limit_rows]:
        body.append(["" if c is None else c for c in r])

    return {
        "headers": headers,
        "rows": body,
        "total_rows": max(0, len(rows) - 1),
        "shown_rows": len(body),
    }


def choose_excel(excel_paths, keyword):
    kw = keyword.lower()
    for p in excel_paths:
        if kw in os.path.basename(p).lower():
            return p
    return None


def parse_epoch_metrics(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {k: [] for k in (
            "epoch","loss","precision","recall","f1","acc","pr_auc_ap","roc_auc","pos_pred_rate","lr","threshold"
        )}

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {h.lower(): i for i, h in enumerate(header)}

    def col_vals(name):
        idx = col.get(name)
        if idx is None:
            return []
        out = []
        for r in rows[1:]:
            if not r or idx >= len(r):
                out.append(None)
            else:
                out.append(safe_float(r[idx]))
        return out

    epoch_list = []
    eidx = col.get("epoch")
    if eidx is not None:
        for r in rows[1:]:
            v = r[eidx] if r and eidx < len(r) else None
            if v is None:
                break
            epoch_list.append(safe_int(v))

    return {
        "epoch": epoch_list,
        "loss": col_vals("loss"),
        "precision": col_vals("precision"),
        "recall": col_vals("recall"),
        "f1": col_vals("f1"),
        "acc": col_vals("acc"),
        "pr_auc_ap": col_vals("pr_auc_ap"),
        "roc_auc": col_vals("roc_auc"),
        "pos_pred_rate": col_vals("pos_pred_rate"),
        "lr": col_vals("lr"),
        "threshold": col_vals("threshold"),
    }


def parse_zip(zip_path):
    data = {
        "zip_path": zip_path,
        "last_modified": os.path.getmtime(zip_path),
        "excel_files": [],
        "model_files": [],
        "train_seconds": None,
        "train": {"excel": None, "run_info": {}, "epoch_metrics": {}, "final_test": {}},
        "test": {"excel": None, "run_info": {}, "final_test": {}, "dataset_summary": None, "per_video": None, "classification_report": None},
    }

    zf = ZipFile(zip_path)
    try:
        excel_paths = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        model_paths = [n for n in zf.namelist() if n.lower().endswith(".pt") or n.lower().endswith(".pth")]
        data["excel_files"] = excel_paths
        data["model_files"] = model_paths

        train_xlsx = choose_excel(excel_paths, "boundary_train_metrics")
        test_xlsx = choose_excel(excel_paths, "cut_test_report")
        data["train"]["excel"] = train_xlsx
        data["test"]["excel"] = test_xlsx

        if train_xlsx:
            with zf.open(train_xlsx) as f:
                wb = load_workbook(filename=f, read_only=True, data_only=True)
                try:
                    if "run_info" in wb.sheetnames:
                        info = parse_kv_sheet(wb["run_info"])
                        data["train"]["run_info"] = info
                        data["train_seconds"] = safe_float(info.get("train_seconds"))
                    if "epoch_metrics" in wb.sheetnames:
                        data["train"]["epoch_metrics"] = parse_epoch_metrics(wb["epoch_metrics"])
                    if "final_test" in wb.sheetnames:
                        data["train"]["final_test"] = parse_kv_sheet(wb["final_test"])
                finally:
                    wb.close()

        if test_xlsx:
            with zf.open(test_xlsx) as f:
                wb = load_workbook(filename=f, read_only=True, data_only=True)
                try:
                    if "run_info" in wb.sheetnames:
                        data["test"]["run_info"] = parse_kv_sheet(wb["run_info"])
                    if "final_test" in wb.sheetnames:
                        data["test"]["final_test"] = parse_kv_sheet(wb["final_test"])
                    if "dataset_summary" in wb.sheetnames:
                        data["test"]["dataset_summary"] = sheet_to_table(wb["dataset_summary"])
                    if "per_video" in wb.sheetnames:
                        data["test"]["per_video"] = sheet_to_table(wb["per_video"])
                    if "classification_report" in wb.sheetnames:
                        data["test"]["classification_report"] = sheet_to_table(wb["classification_report"])
                finally:
                    wb.close()
    finally:
        zf.close()

    return data
